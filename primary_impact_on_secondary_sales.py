#!/usr/bin/env python
# coding: utf-8

# import
import glob
import pandas as pd
import duckdb
import openpyxl
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

# read SKUs
def read_sku_dates():
    file = r"C:\Users\Shithi.Maitra\Downloads\CCF Files\Constraints Packs MTD June 23.xlsx"
    df1 = pd.read_excel(open(file, "rb"), sheet_name="Constraint (Manual Restrict)", header=1, index_col=None)
    df1 = duckdb.query('''select "Base Pack" Basepack, left("Start", 10)::date strt_date, left("End", 10)::date end_date, 'constraint' modality from df1''').df()
    df2 = pd.read_excel(open(file, "rb"), sheet_name="Removed from Order Page ", header=1, index_col=None)
    df2 = duckdb.query('''select "BP Name" Basepack, left("Start", 10)::date strt_date, left("End", 10)::date end_date, 'removed' modality from df2''').df()
    sku_date_df = duckdb.query('''select * from df1 union select * from df2''').df()
    return sku_date_df

# read SCCF
def read_sccf(): 
    
    # files
    path = r"C:\Users\Shithi.Maitra\Downloads\CCF Files"
    filenames = glob.glob(path + r'\*CCFOT*.xlsx')
    
    # read
    sccf_df = pd.DataFrame()
    for f in filenames:
        print("Reading: " + f)
        df = pd.read_excel(open(f, "rb"), sheet_name="Sheet1", header=0, index_col=None)
        df = df[['Orig.Req.Delivery Date', 'Pack Size', 'Final\nOrder Qty.', 'Invoiced\nQuantity']]
        df.columns = ['report_date', 'Basepack', 'sec_ord_qty', 'sec_inv_qty']
        df = duckdb.query('''select strptime(report_date, '%d.%m.%Y')::date report_date, Basepack, sec_ord_qty, sec_inv_qty from df''').df()
        sccf_df = sccf_df.append(df)
    return sccf_df

# read data
sku_dates_df = read_sku_dates()
sccf_read_df = read_sccf()

# read primary data
prim_df = pd.read_excel(open(r"C:\Users\Shithi.Maitra\Downloads\CCF Files\Primary Linefill_YTD 2023.xlsx", "rb"), sheet_name="UBL&UCL", header=0, index_col=None)
prim_df.columns = ['report_date', 'Basepack', 'prim_ord_qty', 'prim_inv_qty']
prim_df = duckdb.query('''select strptime(report_date, '%d.%m.%Y')::date report_date, Basepack, prim_ord_qty, prim_inv_qty from prim_df''').df()
prim_df = duckdb.query('''select * from prim_df where report_date>='2023-01-01' and report_date<'2023-07-01' ''').df()
prim_df

# correct date - delete later
sku_dates_df = read_sku_dates()

# contribution
qry = '''
select Basepack, sec_ord_qty/tot_ord_qty Contribution
from
    (select Basepack, sum(sec_ord_qty) sec_ord_qty
    from sccf_read_df 
    group by 1
    ) tbl1, 

    (select sum(sec_ord_qty) tot_ord_qty
    from sccf_read_df
    ) tbl2
''' 
contrib_df = duckdb.query(qry).df()
contrib_df

# selective data
qry = '''
select distinct tbl1.*, status
from
    (select * 
    from
        (select left(report_date, 10) report_date, Basepack, sum(prim_ord_qty) prim_ord_qty, sum(prim_inv_qty) prim_inv_qty 
        from 
            prim_df tbl1
            inner join 
            (select distinct Basepack from sku_dates_df) tbl2 using(Basepack)
        group by 1, 2
        ) tbl0

        full join

        (select left(report_date, 10) report_date, Basepack, sum(sec_ord_qty) sec_ord_qty, sum(sec_inv_qty) sec_inv_qty 
        from 
            sccf_read_df tbl1
            inner join 
            (select distinct Basepack from sku_dates_df) tbl2 using(Basepack)
        group by 1, 2
        ) tbl1 using(report_date, Basepack)
    ) tbl1
    
    left join 
    
    (select Basepack, strt_date, end_date, max(modality) status
    from sku_dates_df
    group by 1, 2, 3
    ) tbl2 on(report_date>=strt_date and report_date<=end_date and tbl1.Basepack=tbl2.Basepack)
'''
sccf_df = duckdb.query(qry).df()
sccf_df

# bring attributes
qry = '''
-- daily SCCF
select report_date, Basepack, '1. Secondary Order Qty' Measure, sec_ord_qty val from sccf_df
union all
select report_date, Basepack, '2. Secondary Invoiced Qty' Measure, sec_inv_qty val from sccf_df
union all
select report_date, Basepack, '3. SCCF' Measure, sec_inv_qty*1.00/sec_ord_qty val from sccf_df

-- daily DR
union all
select report_date, Basepack, '4. Primary Order Qty' Measure, prim_ord_qty val from sccf_df
union all
select report_date, Basepack, '5. Dispatched Qty' Measure, prim_inv_qty val from sccf_df
union all
select report_date, Basepack, '6. DR' Measure, prim_inv_qty*1.00/prim_ord_qty val from sccf_df

-- Overall SCCF
union all
select 'Overall SCCF' report_date, Basepack, '3. SCCF' Measure, sum(sec_inv_qty)*1.00/sum(sec_ord_qty) val 
from sccf_df
group by 1, 2, 3

-- SCCF - Constrainted/Removed
union all
select 'SCCF - Constrainted/Removed' report_date, Basepack, '3. SCCF' Measure, sum(sec_inv_qty)*1.00/sum(sec_ord_qty) val 
from sccf_df
where status is not null
group by 1, 2, 3

-- SCCF - not Constrainted/Removed
union all
select 'SCCF - not Constrainted/Removed' report_date, Basepack, '3. SCCF' Measure, sum(sec_inv_qty)*1.00/sum(sec_ord_qty) val 
from sccf_df
where status is null
group by 1, 2, 3

-- modality
union all
select 
    report_date, Basepack, '7. Primary Status' Measure, 
    case 
        when status='constraint' then -1 
        when status='removed' then -2
        when status is null and prim_ord_qty is not null then -3
        when status is null and prim_ord_qty is null then -4
    end val 
from sccf_df
'''
piv_sccf_df = duckdb.query(qry).df()

# add contribution
qry = '''
select * 
from 
    piv_sccf_df tbl1 
    inner join 
    contrib_df tbl2 using(Basepack)
'''
piv_sccf_df = duckdb.query(qry).df()
piv_sccf_df

# pivot
piv_df = pd.pivot_table(piv_sccf_df, values='val', index=['Contribution', 'Basepack', 'Measure'], columns='report_date', sort=True)
display(piv_df)

# cosmetic
piv_df = piv_df.replace(-1, "constrainted")
piv_df = piv_df.replace(-2, "removed")
piv_df = piv_df.replace(-3, "on")
piv_df = piv_df.replace(-4, "no data")
piv_df = piv_df.replace(-4, "on")
display(piv_df)

# analyses
qry = '''
select
    Contribution,
    
    Basepack, 
    
    sum(sec_ord_qty) sec_ord_qty, 
    sum(sec_inv_qty) sec_inv_qty, 
    sum(sec_inv_qty)/sum(sec_ord_qty) SCCF, 
    
    sum(case when status is not null then sec_ord_qty else null end) sec_ord_qty_consrem, 
    sum(case when status is not null then sec_inv_qty else null end) sec_inv_qty_consrem, 
    sum(case when status is not null then sec_inv_qty else null end)/sum(case when status is not null then sec_ord_qty else null end) SCCF_consrem, 
    
    sum(case when status='constraint' then sec_ord_qty else null end) sec_ord_qty_cons, 
    sum(case when status='constraint' then sec_inv_qty else null end) sec_inv_qty_cons, 
    sum(case when status='constraint' then sec_inv_qty else null end)/sum(case when status='constraint' then sec_ord_qty else null end) SCCF_cons, 
    
    sum(case when status='removed' then sec_ord_qty else null end) sec_ord_qty_rem, 
    sum(case when status='removed' then sec_inv_qty else null end) sec_inv_qty_rem, 
    sum(case when status='removed' then sec_inv_qty else null end)/sum(case when status='removed' then sec_ord_qty else null end) SCCF_rem, 
    
    sum(case when status is null then sec_ord_qty else null end) sec_ord_qty_not_consrem, 
    sum(case when status is null then sec_inv_qty else null end) sec_inv_qty_not_consrem, 
    sum(case when status is null then sec_inv_qty else null end)/sum(case when status is null then sec_ord_qty else null end) SCCF_not_consrem,

    count(case when status='constraint' then Basepack else null end) days_cons, 
    count(case when status='removed' then Basepack else null end) days_rem, 
    count(case when status is not null then Basepack else null end) days_consrem
from 
    sccf_df tbl1
    inner join 
    contrib_df tbl2 using(Basepack)
group by 1, 2
order by 1 desc
'''
summ_df = duckdb.query(qry).df()
display(summ_df)

# write
path = "primary_off_impact.xlsx"
writer = pd.ExcelWriter(path, engine = 'openpyxl')
piv_df.to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=0, index=True)
summ_df.to_excel(writer, sheet_name='Sheet2', startrow=1, startcol=0, index=False)
writer.close()

# format
workbook = load_workbook(path)
worksheet = workbook["Sheet1"]

# color - red
red_text = Font(color="9C0006")
red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(font=red_text, fill=red_fill)
rule = Rule(type="containsText", operator="containsText", text="constrainted", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("constrainted", D2)))']
worksheet.conditional_formatting.add('D2:GZ2000', rule)

# color - red
red_text = Font(color="9C0006")
red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(font=red_text, fill=red_fill)
rule = Rule(type="containsText", operator="containsText", text="removed", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("removed", D2)))']
worksheet.conditional_formatting.add('D2:GZ2000', rule)

# color - green
green_text = Font(color="006100")
green_fill = PatternFill(bgColor="C6EFCE")
dxf = DifferentialStyle(font=green_text, fill=green_fill)
rule = Rule(type="containsText", operator="containsText", text="on", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("on", D2)))']
worksheet.conditional_formatting.add('D2:GZ2000', rule)

# color - yelllow
yellow_text = Font(color="CC7722")
yellow_fill = PatternFill(bgColor="FFFF00")
dxf = DifferentialStyle(font=yellow_text, fill=yellow_fill)
rule = Rule(type="containsText", operator="containsText", text="no data", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("no data", D2)))']
worksheet.conditional_formatting.add('D2:GZ2000', rule)

# percent
column_letters = [col.column_letter for col in worksheet[1]]
for col in column_letters: 
    for cel in worksheet[(col)]: 
        cel.number_format = "0.0000" 

# freeze
worksheet.freeze_panes = worksheet['D2']

# save, close
workbook.save(path)
workbook.close()

# test an SKU
qry = '''select * from sccf_df where Basepack='VASELINE BODY PETRLUM JELLY (3X9ML+50ML)' order by 1 '''
df = duckdb.query(qry).df()
display(df)




